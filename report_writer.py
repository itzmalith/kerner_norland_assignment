import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from utils import log_step, current_date_str
import pandas as pd
from typing import Dict


@log_step
def style_sheet(sheet, df: pd.DataFrame, bold=True, header_fill_color="B7DEE8"):
    '''
        Applies formatting to an Excel sheet

        :param sheet: openpyxl worksheet to style.
        :param df: DataFrame corresponding to this sheet.
        :param bold: Whether to make header text bold.
        :param header_fill_color: Fill color for header row.
        :return: None
    '''
    bold_font = Font(bold=bold)
    fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center")

    # Stylin header row
    for col_idx in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = bold_font
        cell.fill = fill
        cell.alignment = alignment

    # Auto adjusting column widths
    for i, col in enumerate(sheet.columns, start=1):
        max_length = max((len(str(c.value)) for c in col if c.value), default=0)
        sheet.column_dimensions[get_column_letter(i)].width = max_length + 2


@log_step
def write_report(summary_df: pd.DataFrame, account_data_dict: Dict[str, pd.DataFrame], output_path: str):
    '''
     Writes the summary and per-account data to an Excel report
     Adds styling, Doc Ageing formula, and totals row

     :param summary_df: DataFrame containing summarized data
     :param account_data_dict: Dictionary mapping account to its DataFrame
     :param output_path: Path where the final Excel file will be saved
     :return: None
    '''
    from openpyxl import load_workbook

    # Write raw data first
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        summary_df.to_excel(writer, sheet_name='Summary', index=False, startrow=3)
        for acc, df in account_data_dict.items():
            df.to_excel(writer, sheet_name=str(acc), index=False)

    # Load workbook for styling
    workbook = load_workbook(output_path)

    # Style per-account sheets
    for acc, df in account_data_dict.items():
        sheet = workbook[str(acc)]
        style_sheet(sheet, df)

        # Add Doc Ageing column
        date_col_idx = list(df.columns).index('Document Date') + 1
        ageing_col_idx = sheet.max_column + 1
        sheet.cell(row=1, column=ageing_col_idx, value='Doc Ageing').font = Font(bold=True)
        for row in range(2, sheet.max_row + 1):
            date_cell_ref = sheet.cell(row=row, column=date_col_idx).coordinate
            sheet.cell(row=row, column=ageing_col_idx, value=f"=TODAY()-{date_cell_ref}")

        # Totals row
        if sheet.max_row >= 2:
            totals_row = sheet.max_row + 1
            sheet.cell(row=totals_row, column=1, value="Total").font = Font(bold=True)

            amt_doc_col_idx = list(df.columns).index('Amount in doc. curr.') + 1
            amt_loc_col_idx = list(df.columns).index('Amount in local currency') + 1
            doc_col_letter = get_column_letter(amt_doc_col_idx)
            loc_col_letter = get_column_letter(amt_loc_col_idx)

            sheet.cell(row=totals_row, column=amt_doc_col_idx,
                       value=f"=SUM({doc_col_letter}2:{doc_col_letter}{sheet.max_row - 1})").font = Font(bold=True)
            sheet.cell(row=totals_row, column=amt_loc_col_idx,
                       value=f"=SUM({loc_col_letter}2:{loc_col_letter}{sheet.max_row - 1})").font = Font(bold=True)

    # Style summary sheet
    summary_sheet = workbook['Summary']
    summary_sheet['A1'] = 'Document Ageing Report'
    summary_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=summary_sheet.max_column)
    summary_sheet['A1'].font = Font(size=14, bold=True)
    summary_sheet['A1'].alignment = Alignment(horizontal="center")
    summary_sheet['B2'] = current_date_str()

    style_sheet(summary_sheet, summary_df)
    workbook.save(output_path)
